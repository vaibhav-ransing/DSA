import pandas as pd
from openpyxl import load_workbook

# Define the Excel file location
report_location = 'your_excel_file.xlsx'

# Create your DataFrame for Sheet1
columns = ['Column1', 'Column2', 'Column3']
dataSheet1 = {'Column1': [1, 2, 3], 'Column2': [4, 5, 6], 'Column3': [7, 8, 9]}
df_filtered = pd.DataFrame(dataSheet1, columns=columns)

# Edit a specific cell using the iat method
row_index = 0
column_index = df_filtered.columns.get_loc('ColumnA')  # Get the index of the desired column
new_value = 'UpdatedValue'
df_filtered.iat[row_index, column_index] = new_value

# Create a new ExcelWriter object with a copy of the existing workbook
wb = load_workbook(filename=report_location)
writer = pd.ExcelWriter(report_location, engine='openpyxl')
writer.book = wb

# Save the modified DataFrame back to the "Filtered_Data" sheet using the new ExcelWriter
df_filtered.to_excel(writer, sheet_name='Filtered_Data', index=False)

# Set "Filtered_Data" as the default sheet
default_sheet_name = 'Filtered_Data'
wb.active = wb.sheetnames.index(default_sheet_name)

# Save the changes to the Excel file
writer.save()
