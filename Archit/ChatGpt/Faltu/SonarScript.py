import pandas as pd
from openpyxl import load_workbook

# Define the Excel file location
report_location = 'your_excel_file.xlsx'

# Load the existing Excel file
wb = load_workbook(filename=report_location)

# Specify the sheet name
sheet_name = 'Filtered_Data'

# Define the row and column indices
row_index = 2  # Change this to the desired row index
column_index = 1  # Change this to the desired column index (0-based)

# Access the cell value using openpyxl
ws = wb[sheet_name]
cell_value = ws.cell(row=row_index + 1, column=column_index + 1).value

print("Cell Value:", cell_value)
